from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ErrorCase:
    row_id: int
    reason: str
    error_types: tuple[str, ...]
    action: str
    fixture: bool = True


PRODUCT_ERROR_CASES: tuple[ErrorCase, ...] = (
    ErrorCase(1293, "상품명 어색", ("bad_product_name_fragment",), "product-name validator + final adjudication review"),
    ErrorCase(1113, "NH(농협) 상품명이 ABL생명으로 귀속", ("company_attribution_error",), "local company re-attribution"),
    ErrorCase(1364, "우리아이더(THE)보장보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1363, "우리아이더(THE)보장보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1803, "12일에 출시한 사망보험을 12일 사망보험으로 파싱", ("bad_product_name_fragment",), "date-fragment product-name rejection"),
    ErrorCase(618, "SBS골프 멀티뷰는 보험과 관련 없음", ("non_insurance_product", "sports_broadcast_service_article"), "article eligibility exclusion"),
    ErrorCase(1563, "현대카드/멀티금융사 기사", ("multi_financial_institution_roundup",), "article eligibility exclusion"),
    ErrorCase(1153, "이에 손해보험 접두어 미제거", ("korean_discourse_prefix", "bad_product_name_fragment"), "strip discourse prefix then reject weak generic"),
    ErrorCase(714, "이를 기반으로 안전운전 시보험은 상품명 아님", ("korean_discourse_prefix", "bad_product_name_fragment"), "strip discourse prefix then reject fragment"),
    ErrorCase(840, "이기우는 모델명", ("model_person_name_as_product", "marketing_or_campaign_only"), "model/person product-name guard"),
    ErrorCase(850, "펫보험 기사인데 간편건강보험으로 처리", ("company_attribution_error", "bad_product_name_fragment"), "final adjudication"),
    ErrorCase(910, "한화 시그니처 여성보험인데 DB손해보험 귀속", ("company_attribution_error",), "local company re-attribution"),
    ErrorCase(556, "멘탈/멘털 관련 동일상품", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(557, "멘탈/멘털 관련 동일상품", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(307, "멘탈/멘털 관련 동일상품", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1601, "이후 변액보험 접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "strip discourse prefix then reject weak generic"),
    ErrorCase(1482, "Wealth 컨슈머 팝콘까지 문장조각", ("non_insurance_product", "bad_product_name_fragment"), "article eligibility + product-name rejection"),
    ErrorCase(1820, "먼저보험 접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(1822, "우선보험 접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(1628, "LG헬로비전 관련 상품명 이상", ("subscription_service_not_insurance",), "article eligibility exclusion"),
    ErrorCase(985, "나아가 장기 보장성보험은 상품명 아님", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(670, "엘리스/앨리스 간병보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(753, "엘리스/앨리스 간병보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(676, "월2000원대 보험은 상품명 아님", ("bad_product_name_fragment",), "product-name rejection"),
    ErrorCase(1704, "유병력 펫보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1705, "유병력 펫보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1706, "유병력 펫보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1707, "유병력 펫보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1719, "유병력 펫보험 통합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1720, "이밖에보험 접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(711, "상품명 아님", ("bad_product_name_fragment",), "product-name rejection"),
    ErrorCase(712, "상품명 아님", ("bad_product_name_fragment",), "product-name rejection"),
    ErrorCase(737, "접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(1789, "접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(816, "접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(1460, "상품명 아님", ("bad_product_name_fragment",), "product-name rejection"),
    ErrorCase(1158, "다수보험사 기사", ("multi_company_article",), "article eligibility exclusion"),
    ErrorCase(1622, "다수금융사 기사", ("multi_financial_institution_roundup",), "article eligibility exclusion"),
    ErrorCase(1528, "기사에 한화손해보험 없음", ("company_attribution_error",), "local company validation"),
    ErrorCase(1347, "참기름 관련 비보험상품", ("non_insurance_product",), "article eligibility exclusion"),
    ErrorCase(1348, "따라보험 접두어/문장조각", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(1723, "다수보험사/금융사, 기사상 상품은 현대해상", ("multi_company_article", "company_attribution_error"), "article eligibility + final adjudication"),
    ErrorCase(1551, "KT 구독서비스", ("subscription_service_not_insurance",), "article eligibility exclusion"),
    ErrorCase(1554, "KT 구독서비스", ("subscription_service_not_insurance",), "article eligibility exclusion"),
    ErrorCase(1556, "KT 구독서비스", ("subscription_service_not_insurance",), "article eligibility exclusion"),
    ErrorCase(1264, "결합한보험 접두어 오류", ("korean_discourse_prefix", "bad_product_name_fragment"), "product-name rejection"),
    ErrorCase(310, "결합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(311, "결합 필요", ("bad_product_name_fragment",), "canonical consolidation fixture"),
    ErrorCase(1311, "상품명 오류", ("bad_product_name_fragment",), "product-name rejection/final adjudication"),
    ErrorCase(757, "상품명 오류", ("bad_product_name_fragment",), "product-name rejection/final adjudication"),
)

EXCLUSIVE_ERROR_CASES: tuple[ErrorCase, ...] = (
    ErrorCase(1396, "명칭 이상", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1384, "보험특허권처럼 구체 내용 없는 명칭", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1296, "보험특허권처럼 구체 내용 없는 명칭", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1139, "정체 불명", ("exclusive_right_bad_subject",), "exclusive final adjudication"),
    ErrorCase(1030, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1112, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1134, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1343, "상품명이 아닌 문구/제네럴리 오귀속", ("exclusive_right_bad_subject", "reinsurer_misattribution"), "subject validator + company guard"),
    ErrorCase(1096, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1040, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(980, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(940, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(716, "상품명이 아닌 문구", ("exclusive_right_bad_subject",), "exclusive subject validator"),
    ErrorCase(1390, "획득년월 2027-06 비정상", ("exclusive_right_future_acquired_month",), "future acquired month review"),
    ErrorCase(1260, "하나손해보험인데 제네럴리 표시", ("reinsurer_misattribution", "company_attribution_error"), "exclude reinsurer/foreign branch"),
    ErrorCase(1364, "메리츠화재인데 제네럴리 표시", ("reinsurer_misattribution", "company_attribution_error"), "exclude reinsurer/foreign branch"),
    ErrorCase(671, "삼성화재인데 제네럴리 표시", ("reinsurer_misattribution", "company_attribution_error"), "exclude reinsurer/foreign branch"),
)


def _resolve_input(path_text: str) -> Path:
    path = Path(path_text)
    if path.exists():
        return path
    downloads = Path.home() / "Downloads" / path.name
    if downloads.exists():
        return downloads
    raise FileNotFoundError(path_text)


def _read_sheet(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, Any]] = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[idx]: raw[idx] if idx < len(raw) else None for idx in range(len(headers))}
            rows.append(row)
        return rows
    finally:
        wb.close()


def _index_by_id(rows: list[dict[str, Any]], id_column: str) -> dict[int, dict[str, Any]]:
    indexed: dict[int, dict[str, Any]] = {}
    for row in rows:
        try:
            row_id = int(row.get(id_column))
        except (TypeError, ValueError):
            continue
        indexed[row_id] = row
    return indexed


def _md_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        cleaned = [str(value if value is not None else "").replace("\n", "<br>").replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(cleaned) + " |")
    return "\n".join(lines)


def _product_report_rows(indexed: dict[int, dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for case in PRODUCT_ERROR_CASES:
        row = indexed.get(case.row_id, {})
        rows.append(
            [
                case.row_id,
                ", ".join(case.error_types),
                row.get("상품명"),
                row.get("보험회사"),
                row.get("대표 보종군"),
                row.get("대표 기사 제목") or row.get("source_article_urls"),
                case.reason,
                case.action,
                "Y" if case.fixture else "N",
            ]
        )
    return rows


def _exclusive_report_rows(indexed: dict[int, dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for case in EXCLUSIVE_ERROR_CASES:
        row = indexed.get(case.row_id, {})
        rows.append(
            [
                case.row_id,
                ", ".join(case.error_types),
                row.get("상품/특약/제도명"),
                row.get("보험회사"),
                row.get("획득년월"),
                row.get("대표 기사 제목"),
                case.reason,
                case.action,
                "Y" if case.fixture else "N",
            ]
        )
    return rows


def _write_fixture_json(output_path: Path, product_rows: dict[int, dict[str, Any]], exclusive_rows: dict[int, dict[str, Any]]) -> Path:
    fixture_path = Path("data/exports/extraction_quality_error_fixtures.json")
    fixture_path.parent.mkdir(parents=True, exist_ok=True)
    fixture = {
        "products": [
            {"case": case.__dict__, "row": product_rows.get(case.row_id, {})}
            for case in PRODUCT_ERROR_CASES
        ],
        "exclusive_rights": [
            {"case": case.__dict__, "row": exclusive_rows.get(case.row_id, {})}
            for case in EXCLUSIVE_ERROR_CASES
        ],
        "report": str(output_path),
    }
    fixture_path.write_text(json.dumps(fixture, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return fixture_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--products", required=True)
    parser.add_argument("--exclusive-rights", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    product_path = _resolve_input(args.products)
    exclusive_path = _resolve_input(args.exclusive_rights)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    product_rows = _index_by_id(_read_sheet(product_path, "상품 비교표"), "상품 ID")
    exclusive_rows = _index_by_id(_read_sheet(exclusive_path, "배타적사용권"), "배타적사용권 ID")
    fixture_path = _write_fixture_json(output_path, product_rows, exclusive_rows)

    body = [
        "# Extraction Quality Error Diagnosis",
        "",
        f"- product workbook: `{product_path}`",
        f"- exclusive-right workbook: `{exclusive_path}`",
        f"- fixture JSON: `{fixture_path}`",
        "",
        "## Product Error Cases",
        "",
        _md_table(
            ["ID", "error_type", "상품명", "보험회사", "보종군", "대표 기사/URL", "진단 메모", "필요 조치", "fixture"],
            _product_report_rows(product_rows),
        ),
        "",
        "## Exclusive Right Error Cases",
        "",
        _md_table(
            ["ID", "error_type", "대상명", "보험회사", "획득년월", "대표 기사", "진단 메모", "필요 조치", "fixture"],
            _exclusive_report_rows(exclusive_rows),
        ),
        "",
        "## Required Guardrail Summary",
        "",
        "- Article eligibility: exclude multi-company, multi-financial roundup, non-insurance product/service, campaign/ad-only, subscription, entertainment model, and sports broadcast articles before queue/import.",
        "- Product name quality: strip Korean discourse prefixes, reject weak/generic leftovers and sentence fragments, and keep rejected names as observations only.",
        "- Company attribution: local context wins over raw LLM/query company; reinsurers and foreign branches cannot own active product/exclusive-right rows.",
        "- Final adjudication: compact-context LLM path must return accept/reject/review/reassign/alias-only style decisions, then pass deterministic validators before active save.",
        "- Sales metrics: require product-level evidence near the product name/alias; reject company-wide earnings/revenue metrics.",
    ]
    output_path.write_text("\n".join(body), encoding="utf-8")
    print(json.dumps({"output": str(output_path), "fixture": str(fixture_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
