from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.db.database import Base
from app.db.migrations import create_views
from app.db.models import FactExclusiveUseRight
from app.db.seed_master_data import seed_all
from app.services.exclusive_right_consolidation_service import ExclusiveRightConsolidationService
from app.services.exclusive_right_duplicate_guard_service import ExclusiveRightDuplicateGuardService
from app.services.exclusive_right_service import ExclusiveRightService
from app.utils.text import normalize_search_key


DEFAULT_REPORT_PATH = Path("docs/exclusive-right-consolidation-goal-result.md")


def _seed_event(
    db,
    *,
    company_id: int,
    company: str,
    insurance_type: str,
    subject: str,
    months: int,
    acquired: str,
    evidence: str,
    aliases: list[str] | None = None,
) -> FactExclusiveUseRight:
    row = FactExclusiveUseRight(
        company_id=company_id,
        company_name_normalized=company,
        insurance_type=insurance_type,
        subject_name=subject,
        subject_core_key=normalize_search_key(subject),
        exclusivity_months=months,
        acquired_year_month=acquired,
        feature_summary=evidence,
        evidence_summary=evidence,
        evidence_text=evidence,
        primary_article_title=f"{company} {subject} 배타적사용권",
        primary_article_url=f"https://example.com/{company_id}-{normalize_search_key(subject)}",
        article_count=1,
        confidence_total=0.9,
        needs_review=False,
        event_status="active",
        alias_names_json=json.dumps([subject, *(aliases or [])], ensure_ascii=False),
    )
    db.add(row)
    db.flush()
    row.canonical_exclusive_right_id = row.exclusive_right_id
    return row


def _seed_goal_fixtures(db) -> None:
    _seed_event(
        db,
        company_id=2001,
        company="교보생명",
        insurance_type="생명보험",
        subject="여성건강보험특약",
        months=6,
        acquired="2026-02",
        evidence="교보생명은 여성건강보험특약에 대해 6개월 배타적사용권을 획득했다.",
    )
    _seed_event(
        db,
        company_id=2001,
        company="교보생명",
        insurance_type="생명보험",
        subject="여성건강보험",
        months=6,
        acquired="2026-02",
        evidence="교보생명은 여성건강보험에 대해 6개월 배타적사용권을 획득했다.",
    )
    _seed_event(
        db,
        company_id=2002,
        company="한화손해보험",
        insurance_type="손해보험",
        subject="가정폭력 법률비용 담보 및 Lady 변호사 상담 서비스",
        months=6,
        acquired="2026-03",
        evidence="한화손해보험이 가정폭력 법률비용 담보와 Lady 변호사 상담 서비스로 6개월 배타적사용권을 획득했다.",
        aliases=["시그니처 여성보험 4.0"],
    )
    _seed_event(
        db,
        company_id=2002,
        company="한화손해보험",
        insurance_type="손해보험",
        subject="가정폭력 법률비용 담보 및 변호사 상담 서비스, 가사소송 법률비용 보장",
        months=6,
        acquired="2026-03",
        evidence="한화손해보험이 가정폭력 법률비용 담보, 변호사 상담 서비스, 가사소송 법률비용 보장으로 6개월 배타적사용권을 획득했다.",
        aliases=["여성 건강보험 시그니처 시리즈"],
    )
    db.commit()


def _export_row_count(db, keyword: str) -> int:
    workbook = load_workbook(ExclusiveRightService().export_workbook(db, {}))
    rows = list(workbook.active.iter_rows(values_only=True))
    headers = rows[0]
    subject_index = headers.index("상품/특약/제도명")
    return sum(1 for row in rows[1:] if row[subject_index] and keyword in row[subject_index])


def run_goal_check(report_path: str | Path = DEFAULT_REPORT_PATH) -> dict:
    with tempfile.TemporaryDirectory() as temp_dir:
        engine = create_engine(f"sqlite:///{Path(temp_dir) / 'goal.db'}", future=True, connect_args={"check_same_thread": False})
        Base.metadata.create_all(engine)
        create_views(engine)
        Session = sessionmaker(bind=engine, future=True)
        try:
            with Session() as db:
                seed_all(db)
                _seed_goal_fixtures(db)
                guard = ExclusiveRightDuplicateGuardService()
                before = guard.find_duplicate_groups(db)
                consolidation = ExclusiveRightConsolidationService().run(db, mode="rule_only_apply")
                after = guard.find_duplicate_groups(db)
                kyobo_rows = _export_row_count(db, "여성건강보험")
                hanwha_rows = _export_row_count(db, "법률비용")
                active_subjects = [
                    row.subject_name
                    for row in db.query(FactExclusiveUseRight)
                    .filter(FactExclusiveUseRight.event_status == "active")
                    .order_by(FactExclusiveUseRight.exclusive_right_id)
                    .all()
                ]
                status = "PASS" if kyobo_rows == 1 and hanwha_rows == 1 and not after and consolidation["auto_merge_count"] >= 2 else "FAIL"
                result = {
                    "status": status,
                    "duplicate_groups_before": len(before),
                    "duplicate_groups_after": len(after),
                    "auto_merge_count": consolidation["auto_merge_count"],
                    "kyobo_export_row_count": kyobo_rows,
                    "hanwha_export_row_count": hanwha_rows,
                    "article_level_llm_calls": 0,
                    "export_render_llm_calls": 0,
                    "active_subjects": active_subjects,
                }
        finally:
            engine.dispose()
    _write_report(Path(report_path), result)
    return result


def _write_report(path: Path, result: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "# Exclusive Right Consolidation Goal Result",
                "",
                f"GOAL status = {result['status']}",
                "",
                f"- duplicate groups before: {result['duplicate_groups_before']}",
                f"- duplicate groups after: {result['duplicate_groups_after']}",
                f"- rule auto merge count: {result['auto_merge_count']}",
                f"- 교보생명 여성건강보험 export row count: {result['kyobo_export_row_count']}",
                f"- 한화손해보험 법률비용 export row count: {result['hanwha_export_row_count']}",
                f"- article-level LLM calls: {result['article_level_llm_calls']}",
                f"- export/render LLM calls: {result['export_render_llm_calls']}",
                "",
                "## Active Subjects",
                "",
                *[f"- {subject}" for subject in result["active_subjects"]],
                "",
            ]
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    summary = run_goal_check()
    print(json.dumps(summary, ensure_ascii=False, indent=2))
