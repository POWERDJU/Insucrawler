from __future__ import annotations

from openpyxl import load_workbook

from app.services.dashboard_service import DashboardService
from app.services.product_consolidation_service import ProductConsolidationService
from tests.test_product_duplicate_guard_goal_cases import _seed_goal_duplicates


def _dashboard_request() -> dict:
    return {
        "include_review": True,
        "include_changed_companies": True,
        "include_short_term_insurers": True,
        "include_excluded_policy_products": True,
    }


def _contains(value: object, needle: str) -> bool:
    return needle in str(value or "")


TONTINE = "\ud1a4\ud2f4"
SIGNATURE = "\uc2dc\uadf8\ub2c8\ucc98"
HEALTH_REFUND = "\uac74\uac15\ud658\uae09"
REFUND = "\ud658\uae09"
SURGERY = "\uc804\uc2e0\ub9c8\ucde8\uc218\uc220"
TONTINE_ALIAS = "\ud1a4\ud2f4(Tontine) \uc5f0\uae08"
SIGNATURE_ALIAS = "\uc2dc\uadf8\ub2c8\ucc98 \uc5ec\uc131\ubcf4\ud5d8 4.0"


def test_dashboard_export_dataset_has_one_canonical_row_per_goal_family(db_session):
    _seed_goal_duplicates(db_session)

    ProductConsolidationService().run(db_session, mode="rule_only_apply", target="all", limit=0)

    service = DashboardService()
    products = service._products(db_session, _dashboard_request())
    names = [item["normalized_product_name"] for item in products]

    assert len([name for name in names if _contains(name, TONTINE)]) == 1
    assert len([name for name in names if _contains(name, SIGNATURE) and _contains(name, "4.0")]) == 1
    assert len([name for name in names if _contains(name, HEALTH_REFUND) or _contains(name, REFUND)]) == 1
    assert len([name for name in names if _contains(name, SURGERY)]) == 1
    assert all((item.get("product_status") or "active") != "merged" for item in products)


def test_dashboard_export_workbook_contains_canonical_rows_and_aliases(db_session):
    _seed_goal_duplicates(db_session)

    ProductConsolidationService().run(db_session, mode="rule_only_apply", target="all", limit=0)

    workbook_file = DashboardService().export_comparison_workbook(db_session, _dashboard_request())
    workbook = load_workbook(workbook_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "") for value in rows[0]]
    body = rows[1:]

    assert "canonical_product_id" not in headers
    assert "product_status" not in headers
    assert "상품명 alias 목록" not in headers
    joined_rows = ["\n".join(str(value or "") for value in row) for row in body]

    assert len([row for row in joined_rows if TONTINE in row]) == 1
    assert len([row for row in joined_rows if SIGNATURE in row and "4.0" in row]) == 1
    assert len([row for row in joined_rows if HEALTH_REFUND in row or REFUND in row]) == 1
    assert len([row for row in joined_rows if SURGERY in row]) == 1
    assert all("merged" not in row for row in joined_rows)
