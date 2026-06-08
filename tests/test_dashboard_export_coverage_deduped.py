from __future__ import annotations

from openpyxl import load_workbook

from app.db import repository
from app.services.dashboard_service import DashboardService


def test_dashboard_export_uses_deduped_major_coverages(db_session):
    product = repository.upsert_product(
        db_session,
        {
            "raw_product_name": "수출용 건강보험",
            "normalized_product_name": "수출용 건강보험",
            "company_name": "삼성화재",
            "insurance_type": "손해보험",
            "release_year_month": "2026-01",
            "primary_product_type_code": "HEALTH_COMPREHENSIVE",
            "needs_review": False,
            "confidence_total": 0.95,
        },
    )
    repository.add_major_coverage(db_session, product.product_id, {"coverage_name_raw": "임신지원금", "confidence": 0.5})
    repository.add_major_coverage(db_session, product.product_id, {"coverage_name_raw": "임신지원금 특약", "confidence": 0.95})
    repository.add_major_coverage(db_session, product.product_id, {"coverage_name_raw": "출산지원금", "confidence": 0.8})
    db_session.commit()

    workbook = load_workbook(DashboardService().export_comparison_workbook(db_session, {"include_review": True}))
    values = [str(cell.value or "") for row in workbook[workbook.sheetnames[0]].iter_rows() for cell in row]
    joined = "\n".join(values)

    assert "임신지원금 특약" in joined
    assert "출산지원금" in joined
    assert "임신지원금\n" not in joined
