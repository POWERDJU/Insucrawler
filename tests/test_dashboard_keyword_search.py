from io import BytesIO

from openpyxl import load_workbook

from app.db import repository
from app.db.models import FactArticle
from app.services.dashboard_service import DashboardService
from app.services.ingestion_service import IngestionService
from app.utils.hashing import sha256_text


def _request(**overrides):
    request = {
        "release_year": "전체",
        "release_years": [],
        "release_month": "전체",
        "insurance_type": "전체",
        "company_names": [],
        "product_type_codes": [],
        "classification_mode": "include_secondary",
        "pivot_preset": "custom",
        "custom_rows": ["company_name", "product_type_name"],
        "custom_columns": [],
        "custom_metrics": ["product_count", "article_count"],
        "include_review": False,
        "min_confidence": 0,
        "include_reinsurers": False,
        "include_foreign_branches": False,
        "include_changed_companies": True,
        "include_short_term_insurers": True,
    }
    request.update(overrides)
    return request


def _seed_keyword_product(db):
    product = IngestionService().upsert_structured_product(
        db,
        {
            "product": {
                "raw_product_name": "LG유플러스 키즈폰 고객 전용 미니 보험",
                "normalized_product_name": "키즈폰 어린이 미니보험",
                "company_name": "삼성화재",
                "insurance_type": "손해보험",
                "release_year_month": "2026-01",
                "primary_product_type_code": "CHILD_ADULT_CHILD",
                "confidence_total": 0.92,
                "needs_review": False,
                "partner_company_name": "LG유플러스",
                "partner_context_summary": "LG유플러스 키즈폰 고객을 위한 어린이 특화 보험",
            },
            "product_type_assignments": [
                {"product_type_code": "CHILD_ADULT_CHILD", "assignment_role": "primary", "confidence": 0.92},
            ],
            "narrative_insights": {
                "feature_summary": "LG유플러스 키즈폰 고객을 위한 어린이 특화 보험",
                "coverage_summary": "어린이 상해 및 생활 위험 보장",
                "confidence": 0.9,
                "needs_review": False,
            },
            "major_coverages": [
                {
                    "coverage_name_raw": "어린이 상해 보장",
                    "coverage_name_normalized": "어린이 상해 보장",
                    "risk_area": "상해",
                    "benefit_type": "unknown",
                    "coverage_summary": "어린이 상해 및 생활 위험 보장",
                    "detail_level": "coverage_group",
                    "confidence": 0.9,
                }
            ],
        },
    )
    article = FactArticle(
        source_api="naver",
        title="LG유플러스, 키즈폰 고객 전용 미니 보험 출시",
        description="키즈폰 이용 고객인 어린이를 위한 미니 보험을 선보였다.",
        url="https://example.com/kids-mini",
        original_url="https://example.com/kids-mini",
        content_hash=sha256_text("kids-mini"),
        extraction_status="extracted",
    )
    db.add(article)
    db.flush()
    repository.link_product_article(db, product.product_id, article.article_id, confidence_total=0.9)
    repository.record_product_alias(db, product, "미니 보험", "미니보험", "미니보험", article_id=article.article_id, source_type="weak_mention")
    repository.record_product_alias(db, product, "키즈폰 이용 고객인 어린이를 위한 미니 보험", "키즈폰 이용 고객인 어린이를 위한 미니보험", None, article_id=article.article_id, source_type="descriptive_alias")
    db.commit()
    return product


def test_dashboard_keyword_search_matches_product_alias_summary_coverage_and_article(db_session):
    _seed_keyword_product(db_session)
    service = DashboardService()

    for keyword in ["키즈폰", "어린이", "미니 보험", "미니보험", "상해"]:
        result = service.query(db_session, _request(keyword=keyword))
        assert result["summary"]["product_count"] == 1
        assert result["products"][0]["normalized_product_name"] == "키즈폰 어린이 미니보험"


def test_dashboard_excel_export_uses_same_keyword_filter_and_alias_columns(db_session):
    product = _seed_keyword_product(db_session)
    other = IngestionService().upsert_structured_product(
        db_session,
        {
            "product": {
                "raw_product_name": "일반 건강보험",
                "normalized_product_name": "일반 건강보험",
                "company_name": "삼성화재",
                "insurance_type": "손해보험",
                "release_year_month": "2026-01",
                "primary_product_type_code": "HEALTH_COMPREHENSIVE",
                "confidence_total": 0.9,
                "needs_review": False,
            },
            "product_type_assignments": [
                {"product_type_code": "HEALTH_COMPREHENSIVE", "assignment_role": "primary", "confidence": 0.9},
            ],
        },
    )

    assert product.product_id != other.product_id
    workbook_file = DashboardService().export_comparison_workbook(db_session, _request(keyword="키즈폰"))
    workbook = load_workbook(BytesIO(workbook_file.getvalue()))
    sheet = workbook["상품 비교표"]
    headers = [cell.value for cell in sheet[1]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

    assert "상품명 alias 목록" in headers
    assert "canonical_product_id" in headers
    assert len(rows) == 1
    row = dict(zip(headers, rows[0]))
    assert row["상품명"] == "키즈폰 어린이 미니보험"
    assert "미니 보험" in (row["상품명 alias 목록"] or "")

    filter_sheet = workbook["적용 필터"]
    filter_rows = {row[0].value: row[1].value for row in filter_sheet.iter_rows(min_row=2)}
    assert filter_rows["keyword"] == "키즈폰"
