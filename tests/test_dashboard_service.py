from app.db.models import DimProduct, FactProductTypeAssignment
from app.schemas.dashboard import DashboardQueryRequest
from app.services.dashboard_service import DashboardService
from app.services.ingestion_service import IngestionService
from openpyxl import load_workbook


def seed_dashboard_product(
    db,
    release_year_month="2026-05",
    name="대시보드 간편 암보험",
    product_type_code="CANCER",
    secondary_product_type_code="SIMPLIFIED_IMPAIRED",
):
    IngestionService().upsert_structured_product(
        db,
        {
            "product": {
                "raw_product_name": name,
                "normalized_product_name": name,
                "company_name": "삼성생명",
                "insurance_type": "생명보험",
                "release_year_month": release_year_month,
                "primary_product_type_code": product_type_code,
                "confidence_total": 0.9,
                "needs_review": False,
            },
            "product_type_assignments": [
                {"product_type_code": product_type_code, "assignment_role": "primary", "confidence": 0.95},
                {"product_type_code": secondary_product_type_code, "assignment_role": "secondary", "confidence": 0.9},
            ],
            "major_coverages": [
                {"coverage_name_raw": "암진단비", "risk_area": "암", "benefit_type": "진단", "detail_level": "exact_coverage", "evidence_text": "암진단비 최대 1억원", "max_amount_krw": 100000000, "confidence": 0.9}
            ],
            "features": {
                "join_age_min": 20,
                "join_age_max": 70,
                "notification_type": "간편고지",
                "sales_channel": "GA",
                "renewal_type": "비갱신",
                "payment_period": "20년납",
                "coverage_period": "100세",
                "evidence_text": "20세부터 70세까지 가입, 20년납 100세 보장",
                "confidence": 0.92,
            },
            "narrative_insights": {
                "coverage_summary": "암진단비 중심",
                "feature_summary": "간편고지형 암보험",
                "marketing_summary": "유병자 고객 대상",
                "missing_info_summary": "보험료 예시 확인 필요",
                "evidence_text": "암진단비 중심 간편고지 상품",
                "confidence": 0.88,
            },
            "sales_metrics": [
                {"metric_name": "판매건수", "metric_value": 120, "metric_unit": "건", "metric_period": "2026-05", "metric_basis": "신계약", "evidence_text": "판매건수 120건", "confidence": 0.8}
            ],
        },
    )


def seed_special_clause_product(db):
    IngestionService().upsert_structured_product(
        db,
        {
            "product": {
                "raw_product_name": "암진단 특별약관",
                "normalized_product_name": "암진단 특별약관",
                "company_name": "삼성생명",
                "insurance_type": "생명보험",
                "release_year_month": "2026-05",
                "primary_product_type_code": "CANCER",
                "confidence_total": 0.95,
                "needs_review": False,
            },
            "product_type_assignments": [
                {"product_type_code": "CANCER", "assignment_role": "primary", "confidence": 0.95},
            ],
        },
    )


def seed_rider_product(db):
    IngestionService().upsert_structured_product(
        db,
        {
            "product": {
                "raw_product_name": "암진단 특약",
                "normalized_product_name": "암진단 특약",
                "company_name": "삼성생명",
                "insurance_type": "생명보험",
                "release_year_month": "2026-05",
                "primary_product_type_code": "CANCER",
                "confidence_total": 0.95,
                "needs_review": False,
            },
            "product_type_assignments": [
                {"product_type_code": "CANCER", "assignment_role": "primary", "confidence": 0.95},
            ],
        },
    )


def dashboard_request(**overrides):
    request = {
        "release_year": "전체",
        "release_years": [],
        "release_month": "전체",
        "insurance_type": "전체",
        "company_names": [],
        "product_type_codes": [],
        "classification_mode": "include_secondary",
        "pivot_preset": "custom",
        "custom_rows": ["company_name", "product_type_name"],
        "custom_columns": [],
        "custom_metrics": ["product_count", "article_count"],
        "include_review": False,
        "min_confidence": 0,
        "include_reinsurers": False,
        "include_foreign_branches": False,
        "include_changed_companies": True,
        "include_short_term_insurers": True,
    }
    request.update(overrides)
    return request


def test_dashboard_options_include_master_values(db_session):
    options = DashboardService().options(db_session)
    assert "전체" in options["years"]
    assert options["years"] == ["전체", "2023", "2024", "2025", "2026"]
    assert {"code": "CANCER", "name": "암보험"} in options["product_types"]
    assert any(item["code"] == "company_product_type" for item in options["pivot_presets"])


def test_dashboard_query_request_defaults_to_simplified_dashboard_mode():
    request = DashboardQueryRequest()
    assert request.classification_mode == "include_secondary"
    assert request.pivot_preset == "custom"
    assert request.release_years == []
    assert request.custom_columns == []
    assert request.include_changed_companies is True
    assert request.include_short_term_insurers is True


def test_dashboard_query_returns_pivot_and_products(db_session):
    seed_dashboard_product(db_session)
    result = DashboardService().query(
        db_session,
        dashboard_request(
            release_year="2026",
            release_month="05",
            insurance_type="생명보험",
            company_names=["삼성생명"],
            product_type_codes=["CANCER"],
        ),
    )
    assert result["summary"]["product_count"] == 1
    assert result["summary"]["coverage_count"] == 1
    assert result["products"][0]["normalized_product_name"] == "대시보드 간편 암보험"
    assert result["pivot_result"]["records"][0]["product_count"] == 1


def test_dashboard_custom_columns_empty_and_rows_fallback(db_session):
    seed_dashboard_product(db_session)
    result = DashboardService().query(
        db_session,
        dashboard_request(custom_rows=[], custom_columns=[], custom_metrics=[]),
    )

    assert result["summary"]["product_count"] == 1
    assert result["pivot_result"]["base"] == "product"
    assert result["pivot_result"]["columns"] == []
    assert result["pivot_result"]["rows"] == ["company_name", "product_type_name"]
    assert {row["product_type_name"] for row in result["pivot_result"]["records"]} >= {"암보험", "간편(유병자)"}


def test_dashboard_coverage_rows_select_coverage_base(db_session):
    seed_dashboard_product(db_session)
    for row_key in ["risk_area", "benefit_type"]:
        result = DashboardService().query(
            db_session,
            dashboard_request(custom_rows=["product_type_name", row_key], custom_columns=[], custom_metrics=[]),
        )
        assert result["pivot_result"]["base"] == "coverage"
        assert result["pivot_result"]["columns"] == []
        assert result["pivot_result"]["metrics"] == ["product_count", "coverage_count"]
        assert result["pivot_result"]["records"][0]["product_count"] == 1
        assert result["pivot_result"]["records"][0]["coverage_count"] == 1


def test_dashboard_empty_company_and_product_type_filters_mean_all(db_session):
    seed_dashboard_product(db_session)
    result = DashboardService().query(
        db_session,
        dashboard_request(company_names=[], product_type_codes=[]),
    )

    assert result["summary"]["product_count"] == 1
    assert result["products"][0]["normalized_product_name"] == "대시보드 간편 암보험"


def test_dashboard_product_type_filter_applies_without_industry_or_company(db_session):
    seed_dashboard_product(db_session, name="암보험 상품", product_type_code="CANCER")
    seed_dashboard_product(db_session, name="연금저축 상품", product_type_code="ANNUITY_SAVINGS")

    result = DashboardService().query(
        db_session,
        dashboard_request(
            insurance_type="전체",
            company_names=[],
            product_type_codes=["ANNUITY_SAVINGS"],
        ),
    )

    assert result["summary"]["product_count"] == 1
    assert [item["normalized_product_name"] for item in result["products"]] == ["연금저축 상품"]


def test_dashboard_product_type_filter_includes_primary_even_without_assignment(db_session):
    seed_dashboard_product(db_session, name="배정 누락 암보험", product_type_code="CANCER")
    product = db_session.query(DimProduct).filter(DimProduct.normalized_product_name == "배정 누락 암보험").one()
    db_session.query(FactProductTypeAssignment).filter(FactProductTypeAssignment.product_id == product.product_id).delete()
    db_session.commit()

    result = DashboardService().query(
        db_session,
        dashboard_request(
            insurance_type="전체",
            company_names=[],
            product_type_codes=["CANCER"],
            classification_mode="include_secondary",
        ),
    )

    assert result["summary"]["product_count"] == 1
    assert result["products"][0]["normalized_product_name"] == "배정 누락 암보험"


def test_dashboard_product_list_excludes_special_clause_products(db_session):
    seed_dashboard_product(db_session)
    seed_special_clause_product(db_session)
    seed_rider_product(db_session)

    result = DashboardService().query(
        db_session,
        dashboard_request(company_names=[], product_type_codes=[]),
    )

    names = {item["normalized_product_name"] for item in result["products"]}
    assert "암진단 특별약관" not in names
    assert "암진단 특약" not in names
    assert result["summary"]["product_count"] == 1


def test_dashboard_release_years_filter_multiple_years(db_session):
    seed_dashboard_product(db_session)
    result = DashboardService().query(
        db_session,
        dashboard_request(release_year="전체", release_years=["2026"], release_month="전체"),
    )

    assert result["summary"]["product_count"] == 1


def test_dashboard_export_comparison_workbook(db_session):
    seed_dashboard_product(db_session)
    workbook_file = DashboardService().export_comparison_workbook(
        db_session,
        dashboard_request(release_year="2026", company_names=[], product_type_codes=[]),
    )
    workbook = load_workbook(workbook_file)
    sheet = workbook["상품 비교표"]
    headers = [cell.value for cell in sheet[1]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

    assert "보조 보종군" not in headers
    assert "항목 구분" not in headers
    assert "항목명" not in headers
    assert "값" not in headers
    assert "근거/설명" not in headers
    assert "confidence" not in headers
    assert "검수필요" not in headers
    assert "관련 URL" not in headers
    assert "업종" not in headers
    assert {"상품명", "고지유형", "납입기간", "상품특징 요약", "주요보장1 보장명", "판매실적1 항목"} <= set(headers)
    assert len(rows) == 1
    row = dict(zip(headers, rows[0]))
    assert row["상품명"] == "대시보드 간편 암보험"
    assert row["고지유형"] == "간편고지"
    assert row["납입기간"] == "20년납"
    assert row["상품특징 요약"] == "간편고지형 암보험"
    assert row["주요보장1 보장명"] == "암진단비"
    assert row["판매실적1 항목"] == "판매건수"


def test_dashboard_export_uses_same_product_type_filter_without_company(db_session):
    seed_dashboard_product(db_session, name="암보험 엑셀 제외", product_type_code="CANCER")
    seed_dashboard_product(db_session, name="연금저축 엑셀 포함", product_type_code="ANNUITY_SAVINGS")

    workbook_file = DashboardService().export_comparison_workbook(
        db_session,
        dashboard_request(
            insurance_type="전체",
            company_names=[],
            product_type_codes=["ANNUITY_SAVINGS"],
        ),
    )
    workbook = load_workbook(workbook_file)
    sheet = workbook["상품 비교표"]
    headers = [cell.value for cell in sheet[1]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

    assert len(rows) == 1
    row = dict(zip(headers, rows[0]))
    assert row["상품명"] == "연금저축 엑셀 포함"


def test_dashboard_hides_release_month_outside_visible_period_in_screen_and_excel(db_session):
    seed_dashboard_product(db_session, release_year_month="2022-12", name="과거 출시 상품")

    result = DashboardService().query(db_session, dashboard_request())
    assert result["products"][0]["release_year_month"] == ""

    workbook_file = DashboardService().export_comparison_workbook(db_session, dashboard_request())
    workbook = load_workbook(workbook_file)
    sheet = workbook["상품 비교표"]
    headers = [cell.value for cell in sheet[1]]
    row = dict(zip(headers, [cell.value for cell in next(sheet.iter_rows(min_row=2))]))
    assert row["출시년월"] is None
