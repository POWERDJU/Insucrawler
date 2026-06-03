from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.db import repository
from app.schemas.dashboard import DashboardQueryRequest
from app.services.dashboard_service import DashboardService
from app.services.ingestion_service import IngestionService
from app.services.monthly_new_product_service import MonthlyNewProductService
from app.services.search_service import SearchService


def _request(**overrides):
    request = DashboardQueryRequest(
        custom_rows=["company_name", "product_type_name"],
        custom_metrics=["product_count", "article_count"],
    ).model_dump()
    request.update(overrides)
    return request


def _seed_product(db, *, raw_name: str, normalized_name: str, release_year_month: str = "2026-05"):
    return IngestionService().upsert_structured_product(
        db,
        {
            "product": {
                "raw_product_name": raw_name,
                "normalized_product_name": normalized_name,
                "company_name": "삼성화재",
                "insurance_type": "손해보험",
                "release_year_month": release_year_month,
                "primary_product_type_code": "HEALTH_COMPREHENSIVE",
                "confidence_total": 0.95,
                "needs_review": False,
            },
            "product_type_assignments": [
                {"product_type_code": "HEALTH_COMPREHENSIVE", "assignment_role": "primary", "confidence": 0.95},
            ],
        },
    )


def test_dashboard_query_excludes_silson_medical_names_and_aliases(db_session):
    _seed_product(db_session, raw_name="무배당 실손의료보험", normalized_name="무배당 실손의료보험")
    _seed_product(db_session, raw_name="OO 실손의료 보험", normalized_name="건강보험 A")
    alias_product = _seed_product(db_session, raw_name="일반 건강보험", normalized_name="건강보험 B")
    _seed_product(db_session, raw_name="실손보험", normalized_name="실손보험")
    repository.record_product_alias(
        db_session,
        alias_product,
        "실손의료 보험",
        "실손의료보험",
        "실손의료보험",
        source_type="test",
    )
    db_session.commit()

    result = DashboardService().query(db_session, _request())
    names = {item["normalized_product_name"] for item in result["products"]}

    assert "무배당 실손의료보험" not in names
    assert "건강보험 A" not in names
    assert "건강보험 B" not in names
    assert "실손보험" in names

    included = DashboardService().query(db_session, _request(include_excluded_policy_products=True))
    included_names = {item["normalized_product_name"] for item in included["products"]}
    assert {"무배당 실손의료보험", "건강보험 A", "건강보험 B", "실손보험"} <= included_names


def test_dashboard_export_and_monthly_board_exclude_silson_medical_products(db_session):
    _seed_product(db_session, raw_name="무배당 실손의료보험", normalized_name="무배당 실손의료보험")
    visible = _seed_product(db_session, raw_name="여성 건강보험", normalized_name="여성 건강보험")

    workbook_file = DashboardService().export_comparison_workbook(db_session, _request())
    workbook = load_workbook(BytesIO(workbook_file.getvalue()))
    sheet = workbook["상품 비교표"]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    flattened = "\n".join(str(value) for row in rows for value in row if value)

    assert "무배당 실손의료보험" not in flattened
    assert "여성 건강보험" in flattened

    board = MonthlyNewProductService().get_monthly_new_products(db_session, year_month="2026-05", fallback_latest=False)
    assert [item["product_id"] for item in board["items"]] == [visible.product_id]


def test_search_service_excludes_silson_medical_by_default(db_session):
    _seed_product(db_session, raw_name="무배당 실손의료보험", normalized_name="무배당 실손의료보험")
    _seed_product(db_session, raw_name="실손보험", normalized_name="실손보험")

    result = SearchService().search_products(db_session, include_review=False)
    names = {item["normalized_product_name"] for item in result}

    assert "무배당 실손의료보험" not in names
    assert "실손보험" in names
    assert "무배당 실손의료보험" in {
        item["normalized_product_name"]
        for item in SearchService().search_products(db_session, include_review=False, include_excluded_policy_products=True)
    }
