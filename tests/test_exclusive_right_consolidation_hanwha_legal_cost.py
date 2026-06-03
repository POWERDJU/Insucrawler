from __future__ import annotations

import json

from openpyxl import load_workbook

from app.db.models import FactExclusiveUseRight
from app.services.exclusive_right_consolidation_service import ExclusiveRightConsolidationService
from app.services.exclusive_right_service import ExclusiveRightService
from app.utils.text import normalize_search_key


def _event(db_session, subject: str, *, aliases: list[str] | None = None, evidence: str | None = None) -> FactExclusiveUseRight:
    evidence_text = evidence or (
        "한화손해보험이 가정폭력 등으로 인한 법률비용 담보와 Lady 변호사 상담 서비스 등 "
        "법률 관련 상품과 서비스로 6개월 배타적사용권을 획득했다."
    )
    row = FactExclusiveUseRight(
        company_id=2002,
        company_name_normalized="한화손해보험",
        insurance_type="손해보험",
        subject_name=subject,
        subject_core_key=normalize_search_key(subject),
        exclusivity_months=6,
        acquired_year_month="2026-03",
        feature_summary="가정폭력 법률비용 담보와 변호사 상담 서비스를 보장한다.",
        evidence_summary=evidence_text,
        evidence_text=evidence_text,
        primary_article_title="한화손보, 여성 법률비용 담보·변호사 상담 서비스 배타적사용권 획득",
        primary_article_url=f"https://example.com/hanwha-{normalize_search_key(subject)}",
        article_count=1,
        confidence_total=0.9,
        needs_review=False,
        event_status="active",
        alias_names_json=json.dumps([subject, *(aliases or [])], ensure_ascii=False),
    )
    db_session.add(row)
    db_session.flush()
    row.canonical_exclusive_right_id = row.exclusive_right_id
    return row


def _export_legal_rows(db_session) -> list[tuple]:
    workbook = load_workbook(ExclusiveRightService().export_workbook(db_session, {}))
    rows = list(workbook.active.iter_rows(values_only=True))
    headers = rows[0]
    subject_index = headers.index("상품/특약/제도명")
    return [row for row in rows[1:] if row[subject_index] and "법률비용" in row[subject_index]]


def test_hanwha_legal_cost_and_lawyer_service_variants_merge_to_one_event(db_session):
    first = _event(
        db_session,
        "가정폭력 법률비용 담보 및 Lady 변호사 상담 서비스",
        aliases=["시그니처 여성보험 4.0"],
    )
    second = _event(
        db_session,
        "가정폭력 법률비용 담보 및 변호사 상담 서비스, 가사소송 법률비용 보장",
        aliases=["여성 건강보험 시그니처 시리즈"],
        evidence=(
            "한화손해보험이 가정폭력 법률비용 담보, 변호사 상담 서비스, "
            "가사소송 법률비용 보장으로 6개월 배타적사용권을 획득했다."
        ),
    )
    db_session.commit()

    result = ExclusiveRightConsolidationService().run(db_session, mode="rule_only_apply")
    db_session.refresh(first)
    db_session.refresh(second)

    assert result["auto_merge_count"] == 1
    active = db_session.query(FactExclusiveUseRight).filter(FactExclusiveUseRight.event_status == "active").one()
    assert "법률비용" in active.subject_name
    assert "변호사 상담" in active.subject_name
    assert "시그니처 여성보험" not in active.subject_name
    assert "시그니처 여성 건강보험" not in active.subject_name
    aliases = json.loads(active.alias_names_json)
    assert "가정폭력 법률비용 담보 및 Lady 변호사 상담 서비스" in aliases
    assert "가정폭력 법률비용 담보 및 변호사 상담 서비스, 가사소송 법률비용 보장" in aliases
    assert db_session.query(FactExclusiveUseRight).filter(FactExclusiveUseRight.event_status == "merged").count() == 1
    assert len(_export_legal_rows(db_session)) == 1
