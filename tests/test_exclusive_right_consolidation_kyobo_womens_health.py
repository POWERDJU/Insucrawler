from __future__ import annotations

import json

from openpyxl import load_workbook

from app.db.models import FactExclusiveUseRight
from app.services.exclusive_right_consolidation_service import ExclusiveRightConsolidationService
from app.services.exclusive_right_service import ExclusiveRightService
from app.utils.text import normalize_search_key


def _event(db_session, subject: str) -> FactExclusiveUseRight:
    row = FactExclusiveUseRight(
        company_id=2001,
        company_name_normalized="교보생명",
        insurance_type="생명보험",
        subject_name=subject,
        subject_core_key=normalize_search_key(subject),
        exclusivity_months=6,
        acquired_year_month="2026-02",
        feature_summary="여성 건강 위험 보장을 위한 배타적사용권",
        evidence_summary=f"교보생명은 {subject}에 대해 6개월 배타적사용권을 획득했다.",
        evidence_text=f"교보생명은 {subject}에 대해 6개월 배타적사용권을 획득했다.",
        primary_article_title=f"교보생명 {subject} 배타적사용권 획득",
        primary_article_url=f"https://example.com/kyobo-{normalize_search_key(subject)}",
        article_count=1,
        confidence_total=0.9,
        needs_review=False,
        event_status="active",
        alias_names_json=json.dumps([subject], ensure_ascii=False),
    )
    db_session.add(row)
    db_session.flush()
    row.canonical_exclusive_right_id = row.exclusive_right_id
    return row


def _export_subject_rows(db_session) -> list[tuple]:
    workbook = load_workbook(ExclusiveRightService().export_workbook(db_session, {}))
    rows = list(workbook.active.iter_rows(values_only=True))
    headers = rows[0]
    subject_index = headers.index("상품/특약/제도명")
    return [row for row in rows[1:] if row[subject_index] and "여성건강보험" in row[subject_index]]


def test_kyobo_womens_health_insurance_and_rider_merge_to_one_event(db_session):
    first = _event(db_session, "여성건강보험특약")
    second = _event(db_session, "여성건강보험")
    db_session.commit()

    result = ExclusiveRightConsolidationService().run(db_session, mode="rule_only_apply")
    db_session.refresh(first)
    db_session.refresh(second)

    assert result["auto_merge_count"] == 1
    active = db_session.query(FactExclusiveUseRight).filter(FactExclusiveUseRight.event_status == "active").one()
    assert active.subject_name == "여성건강보험특약"
    aliases = json.loads(active.alias_names_json)
    assert "여성건강보험특약" in aliases
    assert "여성건강보험" in aliases
    assert db_session.query(FactExclusiveUseRight).filter(FactExclusiveUseRight.event_status == "merged").count() == 1
    assert len(_export_subject_rows(db_session)) == 1
