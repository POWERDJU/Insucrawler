from __future__ import annotations

from datetime import datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.main import app
from app.db.database import get_db
from app.db.models import FactArticle
from app.services.exclusive_right_service import ExclusiveRightService


def _seed_right(db, *, text: str, title: str, pub_month: str = "2026-05"):
    article = FactArticle(
        source_api="test",
        title=title,
        description=title,
        url=f"https://example.com/{title}",
        original_url=f"https://example.com/original/{title}",
        pub_date=datetime.fromisoformat(f"{pub_month}-05T09:00:00"),
        content_hash=f"exclusive-export-filter-{title}",
    )
    db.add(article)
    db.commit()
    db.refresh(article)
    right = ExclusiveRightService().create_from_text(db, text, article=article)
    assert right is not None
    return right


def test_exclusive_right_export_applies_current_filters(db_session):
    _seed_right(
        db_session,
        text="한화손해보험은 암 신담보에 대해 6개월 배타적사용권을 획득했다.",
        title="한화손해보험 암 신담보",
    )
    _seed_right(
        db_session,
        text="신한라이프는 연금 특약에 대해 3개월 배타적사용권을 획득했다.",
        title="신한라이프 연금 특약",
    )

    def override_db():
        yield db_session

    app.dependency_overrides[get_db] = override_db
    try:
        response = TestClient(app).post(
            "/api/exclusive-rights/export",
            json={
                "insurance_type": "손해보험",
                "acquired_year_month_from": "2026-05",
                "acquired_year_month_to": "2026-05",
                "keyword": "암",
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["배타적사용권"]
    headers = [cell.value for cell in sheet[1]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

    assert headers == [
        "배타적사용권 ID",
        "업종",
        "보험회사",
        "상품/특약/제도명",
        "배타적사용권 기간 개월 수",
        "획득년월",
        "주요 특징",
        "대표 기사 제목",
        "대표 기사 URL",
        "alias 목록",
        "근거문장",
    ]
    assert len(rows) == 1
    row = dict(zip(headers, rows[0]))
    assert row["업종"] == "손해보험"
    assert row["보험회사"] == "한화손해보험"
    assert "암" in (row["대표 기사 제목"] or "")
