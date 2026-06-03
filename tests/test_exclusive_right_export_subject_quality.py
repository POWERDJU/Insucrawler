from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.db.models import FactExclusiveUseRight
from app.services.exclusive_right_service import ExclusiveRightService


def _event(db_session, subject: str, *, event_status: str = "active", needs_review: bool = False) -> FactExclusiveUseRight:
    row = FactExclusiveUseRight(
        company_id=1,
        company_name_normalized="ABL생명",
        insurance_type="생명보험",
        subject_name=subject,
        subject_core_key="".join(ch for ch in subject if ch.isalnum()).casefold(),
        exclusivity_months=9,
        acquired_year_month="2026-01",
        feature_summary=f"{subject} 배타적사용권",
        evidence_summary=f"{subject} 9개월 배타적사용권",
        evidence_text=f"ABL생명은 {subject}에 대해 9개월 배타적사용권을 획득했다.",
        article_count=1,
        confidence_total=0.9,
        needs_review=needs_review,
        event_status=event_status,
        alias_names_json=f'["{subject}"]',
    )
    db_session.add(row)
    db_session.flush()
    row.canonical_exclusive_right_id = row.exclusive_right_id
    return row


def test_exclusive_right_export_excludes_bad_subjects_by_default(db_session):
    _event(db_session, "상품")
    _event(db_session, "보장 특약을 개발해 손해 보험")
    good = _event(db_session, "우리WON건강환급보험")
    db_session.commit()

    workbook = ExclusiveRightService().export_workbook(db_session, {})
    sheet = load_workbook(BytesIO(workbook.getvalue()))["배타적사용권"]
    subjects = [row[3].value for row in sheet.iter_rows(min_row=2)]

    assert subjects == [good.subject_name]
    assert "상품" not in subjects
    assert "보장 특약을 개발해 손해 보험" not in subjects
